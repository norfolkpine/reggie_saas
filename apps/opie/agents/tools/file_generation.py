import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from uuid import uuid4
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from agno.media import File
from agno.tools import Toolkit
from agno.tools.function import ToolResult
from agno.utils.log import log_debug, logger

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed. PDF generation will not be available. Install with: pip install reportlab")

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not installed. XLSX generation will not be available. Install with: pip install openpyxl")


class FileGenerationTools(Toolkit):
    def __init__(
        self,
        enable_json_generation: bool = True,
        enable_csv_generation: bool = True,
        enable_pdf_generation: bool = True,
        enable_txt_generation: bool = True,
        enable_xlsx_generation: bool = True,
        output_directory: Optional[str] = None,
        user_uuid: Optional[str] = None,
        all: bool = False,
        **kwargs,
    ):
        self.enable_json_generation = enable_json_generation
        self.enable_csv_generation = enable_csv_generation
        self.enable_pdf_generation = enable_pdf_generation and PDF_AVAILABLE
        self.enable_txt_generation = enable_txt_generation
        self.enable_xlsx_generation = enable_xlsx_generation and XLSX_AVAILABLE
        self.output_directory = Path(output_directory) if output_directory else None
        self.user_uuid = user_uuid

        # Create output directory if specified (only for local development)
        if self.output_directory and settings.DEBUG:
            self.output_directory.mkdir(parents=True, exist_ok=True)
            log_debug(f"Files will be saved to: {self.output_directory}")

        if enable_pdf_generation and not PDF_AVAILABLE:
            logger.warning("PDF generation requested but reportlab is not installed. Disabling PDF generation.")
            self.enable_pdf_generation = False

        if enable_xlsx_generation and not XLSX_AVAILABLE:
            logger.warning("XLSX generation requested but openpyxl is not installed. Disabling XLSX generation.")
            self.enable_xlsx_generation = False

        tools: List[Any] = []
        if all or enable_json_generation:
            tools.append(self.generate_json_file)
        if all or enable_csv_generation:
            tools.append(self.generate_csv_file)
        if all or (enable_pdf_generation and PDF_AVAILABLE):
            tools.append(self.generate_pdf_file)
        if all or enable_txt_generation:
            tools.append(self.generate_text_file)
        if all or (enable_xlsx_generation and XLSX_AVAILABLE):
            tools.append(self.generate_xlsx_file)

        super().__init__(name="file_generation", tools=tools, **kwargs)

    def _get_partitioned_path(self, filename: str) -> str:
        """
        Generate partitioned path structure: user_files/user_uuid=<uuid>/year=<year>/month=<month>/day=<day>/filename
        """
        now = datetime.now()

        # Use provided user_uuid or generate a default one
        user_id = self.user_uuid if self.user_uuid else "anonymous"

        # Build partitioned path
        path_parts = [
            "user_files",
            f"user_uuid={user_id}",
            f"year={now.year}",
            f"month={now.month:02d}",
            f"day={now.day:02d}",
            filename
        ]

        return "/".join(path_parts)

    def _save_file_to_disk(self, content: Union[str, bytes], filename: str) -> Optional[str]:
        """Save file using Django's storage backend (GCS in production, local in dev). Return URL or None."""
        try:
            # Generate partitioned path structure
            storage_path = self._get_partitioned_path(filename)

            # Convert content to bytes if needed
            if isinstance(content, str):
                content_bytes = content.encode('utf-8')
            else:
                content_bytes = content

            # Save using Django's default storage (automatically uses GCS in production, local in dev)
            saved_path = default_storage.save(storage_path, ContentFile(content_bytes))

            # Get the public URL
            url = default_storage.url(saved_path)

            log_debug(f"File saved to storage: {saved_path}, URL: {url}")
            return url

        except Exception as e:
            logger.error(f"Failed to save file to storage: {e}")
            return None

    def generate_json_file(self, data: Union[Dict, List, str], filename: Optional[str] = None) -> ToolResult:
        """Generate a JSON file from the provided data.

        Args:
            data: The data to write to the JSON file. Can be a dictionary, list, or JSON string.
            filename: Optional filename for the generated file. If not provided, a UUID will be used.

        Returns:
            ToolResult: Result containing the generated JSON file as a FileArtifact.
        """
        try:
            log_debug(f"Generating JSON file with data: {type(data)}")

            # Handle different input types
            if isinstance(data, str):
                try:
                    json.loads(data)
                    json_content = data  # Use the original string if it's valid JSON
                except json.JSONDecodeError:
                    # If it's not valid JSON, treat as plain text and wrap it
                    json_content = json.dumps({"content": data}, indent=2)
            else:
                json_content = json.dumps(data, indent=2, ensure_ascii=False)

            # Generate filename if not provided
            if not filename:
                filename = f"generated_file_{str(uuid4())[:8]}.json"
            elif not filename.endswith(".json"):
                filename += ".json"

            # Save file to disk (if output_directory is set)
            file_path = self._save_file_to_disk(json_content, filename)

            # Create FileArtifact
            file_artifact = File(
                id=str(uuid4()),
                content=json_content,
                mime_type="application/json",
                file_type="json",
                filename=filename,
                size=len(json_content.encode("utf-8")),
                url=f"{file_path}" if file_path else None,
            )

            log_debug("JSON file generated successfully")
            success_msg = f"JSON file '{filename}' has been generated successfully with {len(json_content)} characters."
            if file_path:
                success_msg += f" File saved to: {file_path}"
            else:
                success_msg += " File is available in response."

            return ToolResult(content=success_msg, files=[file_artifact])

        except Exception as e:
            logger.error(f"Failed to generate JSON file: {e}")
            return ToolResult(content=f"Error generating JSON file: {e}")

    def generate_csv_file(
        self,
        data: Union[List[List], List[Dict], str],
        filename: Optional[str] = None,
        headers: Optional[List[str]] = None,
    ) -> ToolResult:
        """Generate a CSV file from the provided data.

        Args:
            data: The data to write to the CSV file. Can be a list of lists, list of dictionaries, or CSV string.
            filename: Optional filename for the generated file. If not provided, a UUID will be used.
            headers: Optional headers for the CSV. Used when data is a list of lists.

        Returns:
            ToolResult: Result containing the generated CSV file as a FileArtifact.
        """
        try:
            log_debug(f"Generating CSV file with data: {type(data)}")

            # Create CSV content
            output = io.StringIO()

            if isinstance(data, str):
                # If it's already a CSV string, use it directly
                csv_content = data
            elif isinstance(data, list) and len(data) > 0:
                writer = csv.writer(output)

                if isinstance(data[0], dict):
                    # List of dictionaries - use keys as headers
                    if data:
                        fieldnames = list(data[0].keys())
                        writer.writerow(fieldnames)
                        for row in data:
                            if isinstance(row, dict):
                                writer.writerow([row.get(field, "") for field in fieldnames])
                            else:
                                writer.writerow([str(row)] + [""] * (len(fieldnames) - 1))
                elif isinstance(data[0], list):
                    # List of lists
                    if headers:
                        writer.writerow(headers)
                    writer.writerows(data)
                else:
                    # List of other types
                    if headers:
                        writer.writerow(headers)
                    for item in data:
                        writer.writerow([str(item)])

                csv_content = output.getvalue()
            else:
                csv_content = ""

            # Generate filename if not provided
            if not filename:
                filename = f"generated_file_{str(uuid4())[:8]}.csv"
            elif not filename.endswith(".csv"):
                filename += ".csv"

            # Save file to disk (if output_directory is set)
            file_path = self._save_file_to_disk(csv_content, filename)

            # Create FileArtifact
            file_artifact = File(
                id=str(uuid4()),
                content=csv_content,
                mime_type="text/csv",
                file_type="csv",
                filename=filename,
                size=len(csv_content.encode("utf-8")),
                url=f"{file_path}" if file_path else None,
            )

            log_debug("CSV file generated successfully")
            success_msg = f"CSV file '{filename}' has been generated successfully with {len(csv_content)} characters."
            if file_path:
                success_msg += f" File saved to: {file_path}"
            else:
                success_msg += " File is available in response."

            return ToolResult(content=success_msg, files=[file_artifact])

        except Exception as e:
            logger.error(f"Failed to generate CSV file: {e}")
            return ToolResult(content=f"Error generating CSV file: {e}")

    def generate_pdf_file(
        self, content: str, filename: Optional[str] = None, title: Optional[str] = None
    ) -> ToolResult:
        """Generate a PDF file from the provided content.

        Args:
            content: The text content to write to the PDF file.
            filename: Optional filename for the generated file. If not provided, a UUID will be used.
            title: Optional title for the PDF document.

        Returns:
            ToolResult: Result containing the generated PDF file as a FileArtifact.
        """
        if not PDF_AVAILABLE:
            return ToolResult(
                content="PDF generation is not available. Please install reportlab: pip install reportlab"
            )

        try:
            log_debug(f"Generating PDF file with content length: {len(content)}")

            # Create PDF content in memory
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1 * inch)

            # Get styles
            styles = getSampleStyleSheet()
            title_style = styles["Title"]
            normal_style = styles["Normal"]

            # Build story (content elements)
            story = []

            if title:
                story.append(Paragraph(title, title_style))
                story.append(Spacer(1, 20))

            # Split content into paragraphs and add to story
            paragraphs = content.split("\n\n")
            for para in paragraphs:
                if para.strip():
                    # Clean the paragraph text for PDF
                    clean_para = para.strip().replace("<", "&lt;").replace(">", "&gt;")
                    story.append(Paragraph(clean_para, normal_style))
                    story.append(Spacer(1, 10))

            # Build PDF
            doc.build(story)
            pdf_content = buffer.getvalue()
            buffer.close()

            # Generate filename if not provided
            if not filename:
                filename = f"generated_file_{str(uuid4())[:8]}.pdf"
            elif not filename.endswith(".pdf"):
                filename += ".pdf"

            # Save file to disk (if output_directory is set)
            file_path = self._save_file_to_disk(pdf_content, filename)

            # Create FileArtifact
            file_artifact = File(
                id=str(uuid4()),
                content=pdf_content,
                mime_type="application/pdf",
                file_type="pdf",
                filename=filename,
                size=len(pdf_content),
                url=f"{file_path}" if file_path else None,
            )

            log_debug("PDF file generated successfully")
            success_msg = f"PDF file '{filename}' has been generated successfully with {len(pdf_content)} bytes."
            if file_path:
                success_msg += f" File saved to: {file_path}"
            else:
                success_msg += " File is available in response."

            return ToolResult(content=success_msg, files=[file_artifact])

        except Exception as e:
            logger.error(f"Failed to generate PDF file: {e}")
            return ToolResult(content=f"Error generating PDF file: {e}")

    def generate_text_file(self, content: str, filename: Optional[str] = None) -> ToolResult:
        """Generate a text file from the provided content.

        Args:
            content: The text content to write to the file.
            filename: Optional filename for the generated file. If not provided, a UUID will be used.

        Returns:
            ToolResult: Result containing the generated text file as a FileArtifact.
        """
        try:
            log_debug(f"Generating text file with content length: {len(content)}")

            # Generate filename if not provided
            if not filename:
                filename = f"generated_file_{str(uuid4())[:8]}.txt"
            elif not filename.endswith(".txt"):
                filename += ".txt"

            # Save file to disk (if output_directory is set)
            file_path = self._save_file_to_disk(content, filename)

            # Create FileArtifact
            file_artifact = File(
                id=str(uuid4()),
                content=content,
                mime_type="text/plain",
                file_type="txt",
                filename=filename,
                size=len(content.encode("utf-8")),
                url=f"{file_path}" if file_path else None,
            )

            log_debug("Text file generated successfully")
            success_msg = f"Text file '{filename}' has been generated successfully with {len(content)} characters."
            if file_path:
                success_msg += f" File saved to: {file_path}"
            else:
                success_msg += " File is available in response."

            return ToolResult(content=success_msg, files=[file_artifact])

        except Exception as e:
            logger.error(f"Failed to generate text file: {e}")
            return ToolResult(content=f"Error generating text file: {e}")

    def generate_xlsx_file(
        self,
        data: Union[List[List], List[Dict], str],
        filename: Optional[str] = None,
        sheet_name: Optional[str] = "Sheet1",
        headers: Optional[List[str]] = None,
    ) -> ToolResult:

        if not XLSX_AVAILABLE:
            return ToolResult(
                content="XLSX generation is not available. Please install openpyxl: pip install openpyxl"
            )

        try:
            log_debug(f"Generating XLSX file with data: {type(data)}")

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            if isinstance(data, str):
                try:
                    import json
                    parsed_data = json.loads(data)
                    if isinstance(parsed_data, list):
                        data = parsed_data
                    else:
                        ws.append([data])
                except json.JSONDecodeError:
                    ws.append([data])
            elif isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], dict):
                    fieldnames = list(data[0].keys())
                    ws.append(fieldnames)
                    for row in data:
                        if isinstance(row, dict):
                            ws.append([row.get(field, "") for field in fieldnames])
                        else:
                            ws.append([str(row)] + [""] * (len(fieldnames) - 1))
                elif isinstance(data[0], list):
                    if headers:
                        ws.append(headers)
                    for row in data:
                        ws.append(row)
                else:
                    if headers:
                        ws.append(headers)
                    for item in data:
                        ws.append([item])

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

            buffer = io.BytesIO()
            wb.save(buffer)
            xlsx_content = buffer.getvalue()
            buffer.close()

            if not filename:
                filename = f"generated_file_{str(uuid4())[:8]}.xlsx"
            elif not filename.endswith(".xlsx"):
                filename += ".xlsx"

            file_path = self._save_file_to_disk(xlsx_content, filename)

            file_artifact = File(
                id=str(uuid4()),
                content=xlsx_content,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                file_type="xlsx",
                filename=filename,
                size=len(xlsx_content),
                url=f"{file_path}" if file_path else None,
            )

            log_debug("XLSX file generated successfully")
            success_msg = f"XLSX file '{filename}' has been generated successfully with {len(xlsx_content)} bytes."
            if file_path:
                success_msg += f" File saved to: {file_path}"
            else:
                success_msg += " File is available in response."

            return ToolResult(content=success_msg, files=[file_artifact])

        except Exception as e:
            logger.error(f"Failed to generate XLSX file: {e}")
            return ToolResult(content=f"Error generating XLSX file: {e}")
